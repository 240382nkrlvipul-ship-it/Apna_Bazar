import io
from flask import Blueprint, send_file, request, jsonify, g
from backend.database import Order, db, Customer, Product, Category, OrderItem
from backend.utils.auth_helper import admin_required
from datetime import datetime, timedelta
from sqlalchemy import func

# reportlab imports for PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# openpyxl imports for Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

reports_bp = Blueprint('reports', __name__)

from backend.utils.cache import cache

@reports_bp.route('/stats', methods=['GET'])
@admin_required
def get_dashboard_stats():
    """Fetches key performance indicators and chart data for admin dashboard"""
    cache_key = "reports:stats"
    cached = cache.get(cache_key)
    if cached is not None:
        return jsonify(cached), 200

    today = datetime.utcnow().date()
    start_of_month = datetime(today.year, today.month, 1)
    today_start = datetime(today.year, today.month, today.day, 0, 0, 0)
    today_end = datetime(today.year, today.month, today.day, 23, 59, 59)
    
    # 1. KPI cards data - optimized via DB aggregates
    total_orders = Order.query.count()
    total_customers = Customer.query.count()
    
    today_sales = db.session.query(func.sum(Order.total_amount)).filter(
        Order.created_at >= today_start,
        Order.created_at <= today_end,
        Order.order_status != 'Cancelled'
    ).scalar() or 0.0
    
    monthly_sales = db.session.query(func.sum(Order.total_amount)).filter(
        Order.created_at >= start_of_month,
        Order.order_status != 'Cancelled'
    ).scalar() or 0.0
    
    pending_orders = Order.query.filter(
        Order.order_status.in_(['Order Received', 'Preparing', 'Packed', 'Assigned To Delivery Agent', 'Out For Delivery'])
    ).count()
    
    delivered_orders = Order.query.filter_by(order_status='Delivered').count()
    low_stock_count = Product.query.filter(Product.stock_quantity <= 5).count()
    
    # 2. Daily revenue chart data (last 7 days)
    revenue_by_day = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_start = datetime(day.year, day.month, day.day, 0, 0, 0)
        day_end = datetime(day.year, day.month, day.day, 23, 59, 59)
        
        day_revenue = db.session.query(func.sum(Order.total_amount)).filter(
            Order.created_at >= day_start,
            Order.created_at <= day_end,
            Order.order_status != 'Cancelled'
        ).scalar() or 0.0
        
        revenue_by_day.append({
            'day': day.strftime('%a'), # E.g., 'Mon', 'Tue'
            'revenue': float(day_revenue)
        })
      
    # 3. Category distribution (revenue by category)
    categories_rev = db.session.query(
        Category.name_en, func.sum(OrderItem.price * OrderItem.quantity)
    ).join(Product, Product.category_id == Category.id)\
     .join(OrderItem, OrderItem.product_id == Product.id)\
     .join(Order, OrderItem.order_id == Order.id)\
     .filter(Order.order_status != 'Cancelled')\
     .group_by(Category.name_en).all()
     
    revenue_by_category = [
        {'category': name, 'revenue': float(rev)}
        for name, rev in categories_rev
    ]
    
    # 4. Top selling products (quantity sold)
    top_prods_query = db.session.query(
        OrderItem.product_name_en, func.sum(OrderItem.quantity)
    ).join(Order, OrderItem.order_id == Order.id)\
     .filter(Order.order_status != 'Cancelled')\
     .group_by(OrderItem.product_name_en)\
     .order_by(func.sum(OrderItem.quantity).desc())\
     .limit(5).all()
     
    top_products = [
        {'name': name, 'sold': float(sold)}
        for name, sold in top_prods_query
    ]
    
    res_dict = {
        'today_sales': float(today_sales),
        'monthly_sales': float(monthly_sales),
        'total_customers': total_customers,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'delivered_orders': delivered_orders,
        'low_stock_count': low_stock_count,
        'revenue_by_day': revenue_by_day,
        'revenue_by_category': revenue_by_category,
        'top_products': top_products
    }
    
    cache.set(cache_key, res_dict, timeout=300)
    return jsonify(res_dict), 200


@reports_bp.route('/pdf', methods=['GET'])
@admin_required
def generate_pdf_report():
    """Generates a downloadable PDF report of grocery store sales"""
    days = request.args.get('days', 30, type=int)
    start_date = datetime.utcnow() - timedelta(days=days)
    
    # Query orders in date range
    orders = Order.query.filter(Order.created_at >= start_date).order_by(Order.created_at.desc()).all()
    
    # Calculate stats
    total_orders = len(orders)
    total_revenue = sum(float(o.total_amount) for o in orders if o.order_status != 'Cancelled')
    pending_orders = sum(1 for o in orders if o.order_status in ['Order Received', 'Preparing', 'Packed', 'Out For Delivery'])
    delivered_orders = sum(1 for o in orders if o.order_status == 'Delivered')
    cancelled_orders = sum(1 for o in orders if o.order_status == 'Cancelled')
    
    # Create file stream
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#0f172a'),
        alignment=1 # Center
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748b'),
        alignment=1 # Center
    )
    
    section_title = ParagraphStyle(
        'SecTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#1e293b'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#334155')
    )
    
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white
    )
    
    # Header Section
    story.append(Paragraph("Village Grocery Store - Sales Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Scope: Past {days} Days", subtitle_style))
    story.append(Spacer(1, 15))
    
    # Summary Statistics Table
    story.append(Paragraph("Executive Summary", section_title))
    summary_data = [
        [
            Paragraph("<b>Total Sales:</b>", body_style), Paragraph(f"₹{total_revenue:,.2f}", body_style),
            Paragraph("<b>Total Orders:</b>", body_style), Paragraph(str(total_orders), body_style)
        ],
        [
            Paragraph("<b>Delivered:</b>", body_style), Paragraph(str(delivered_orders), body_style),
            Paragraph("<b>Pending:</b>", body_style), Paragraph(str(pending_orders), body_style)
        ],
        [
            Paragraph("<b>Cancelled:</b>", body_style), Paragraph(str(cancelled_orders), body_style),
            Paragraph("", body_style), Paragraph("", body_style)
        ]
    ]
    summary_table = Table(summary_data, colWidths=[120, 140, 120, 140])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 15))
    
    # Orders Detail Table
    story.append(Paragraph("Recent Orders Log", section_title))
    orders_headers = [
        Paragraph("Order ID", header_style),
        Paragraph("Customer Name", header_style),
        Paragraph("Mobile", header_style),
        Paragraph("Payment", header_style),
        Paragraph("Status", header_style),
        Paragraph("Amount", header_style),
        Paragraph("Date", header_style)
    ]
    
    table_data = [orders_headers]
    
    for order in orders:
        created_str = order.created_at.strftime('%Y-%m-%d') if order.created_at else ''
        table_data.append([
            Paragraph(f"#{order.id}", body_style),
            Paragraph(order.customer_name, body_style),
            Paragraph(order.customer_mobile, body_style),
            Paragraph(order.payment_method, body_style),
            Paragraph(order.order_status, body_style),
            Paragraph(f"₹{float(order.total_amount):.2f}", body_style),
            Paragraph(created_str, body_style)
        ])
        
    orders_table = Table(table_data, colWidths=[55, 110, 80, 55, 95, 65, 70])
    orders_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(orders_table)
    
    # Build Document
    doc.build(story)
    
    buffer.seek(0)
    filename = f"sales_report_{days}_days.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@reports_bp.route('/excel', methods=['GET'])
@admin_required
def generate_excel_report():
    """Generates a downloadable Excel (.xlsx) file of grocery store sales"""
    days = request.args.get('days', 30, type=int)
    start_date = datetime.utcnow() - timedelta(days=days)
    
    orders = Order.query.filter(Order.created_at >= start_date).order_by(Order.created_at.desc()).all()
    
    # Initialize openpyxl workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Summary Dashboard
    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    ws1.views.sheetView[0].showGridLines = True
    
    # Custom colors and styles
    navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    accent_fill = PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid")
    
    font_title = Font(name="Arial", size=16, bold=True, color="0F172A")
    font_section = Font(name="Arial", size=12, bold=True, color="1E293B")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)
    
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Title Block
    ws1['A1'] = "Village Grocery Store - Sales Report Dashboard"
    ws1['A1'].font = font_title
    ws1['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Range: Past {days} Days"
    ws1['A2'].font = Font(name="Arial", size=10, italic=True, color="64748B")
    
    # Statistics Section
    ws1['A4'] = "KPI Metrics"
    ws1['A4'].font = font_section
    
    ws1['A5'] = "Metric"
    ws1['B5'] = "Value"
    ws1['A5'].font = font_header
    ws1['B5'].font = font_header
    ws1['A5'].fill = navy_fill
    ws1['B5'].fill = navy_fill
    ws1['A5'].border = border_thin
    ws1['B5'].border = border_thin
    
    total_sales = sum(float(o.total_amount) for o in orders if o.order_status != 'Cancelled')
    delivered = sum(1 for o in orders if o.order_status == 'Delivered')
    pending = sum(1 for o in orders if o.order_status in ['Order Received', 'Preparing', 'Packed', 'Out For Delivery'])
    cancelled = sum(1 for o in orders if o.order_status == 'Cancelled')
    
    metrics = [
        ("Total Revenue (Excl. Cancelled)", f"INR {total_sales:,.2f}"),
        ("Total Orders Placed", len(orders)),
        ("Delivered Orders", delivered),
        ("Pending Orders", pending),
        ("Cancelled Orders", cancelled)
    ]
    
    row_idx = 6
    for key, val in metrics:
        ws1.cell(row=row_idx, column=1, value=key).font = font_bold
        ws1.cell(row=row_idx, column=1).border = border_thin
        ws1.cell(row=row_idx, column=1).fill = light_blue_fill
        
        c_val = ws1.cell(row=row_idx, column=2, value=val)
        c_val.font = font_regular
        c_val.border = border_thin
        c_val.alignment = Alignment(horizontal="right")
        row_idx += 1
        
    # Auto-adjust column width for Sheet 1
    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 20
    
    # Sheet 2: Detailed Orders Log
    ws2 = wb.create_sheet(title="Detailed Orders Log")
    ws2.views.sheetView[0].showGridLines = True
    
    headers = [
        "Order ID", "Customer Name", "Customer Mobile", 
        "Delivery Address", "Village ID", "Payment Method", 
        "Payment Status", "Order Status", "Delivery Charge", 
        "Total Amount (INR)", "Date Placed"
    ]
    
    # Write Headers
    for col_idx, text in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
        
    # Write Data
    for row_num, order in enumerate(orders, 2):
        created_str = order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else ''
        ws2.cell(row=row_num, column=1, value=f"#{order.id}")
        ws2.cell(row=row_num, column=2, value=order.customer_name)
        ws2.cell(row=row_num, column=3, value=order.customer_mobile)
        ws2.cell(row=row_num, column=4, value=order.delivery_address)
        ws2.cell(row=row_num, column=5, value=order.village_id)
        ws2.cell(row=row_num, column=6, value=order.payment_method)
        ws2.cell(row=row_num, column=7, value=order.payment_status)
        ws2.cell(row=row_num, column=8, value=order.order_status)
        ws2.cell(row=row_num, column=9, value=float(order.delivery_charge))
        ws2.cell(row=row_num, column=10, value=float(order.total_amount))
        ws2.cell(row=row_num, column=11, value=created_str)
        
        # Apply style to data rows
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(row=row_num, column=col_idx)
            cell.font = font_regular
            cell.border = border_thin
            if col_idx in [1, 3, 5, 11]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [9, 10]:
                cell.alignment = Alignment(horizontal="right")
                
    # Auto-adjust column widths for Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Write workbook to stream
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"sales_report_{days}_days.xlsx"
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
